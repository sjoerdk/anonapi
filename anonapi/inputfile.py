"""Classes and functions Reading job parameters from csv and excel files

These files contain tabular data with identifiers such as accession numbers
and optionally pseudonyms.
It is often easier to add such a file to a mapping programatically then to
copy-paste between open files
"""
import csv
import logging
from pathlib import Path
from typing import Iterator, List, Type, Union, Optional

from openpyxl.reader.excel import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from anonapi.exceptions import AnonAPIException
from anonapi.mapper import JobParameterGrid, sniff_dialect_safe
from anonapi.parameters import (
    AccessionNumber,
    Parameter,
    ParameterFactory,
    ParameterParsingError,
    PathParameter,
    PseudoName,
)


logger = logging.getLogger(__name__)


class ParameterColumn:
    """A column of Parameter instances, like accession numbers or pseudonyms
    supposed to be part of a grid like a csv file or xls file

    Does fuzzy finding of column header and parsing of values
    """

    header_names: List[str] = []  # the column_types that could be above this column

    # the type of parameter that this column contains
    parameter_type: Type[Parameter] = Parameter

    def __init__(self, column: int, header_row_idx: int = 0):
        """

        Parameters
        ----------
        column: int
            The 0-based column in which this header instance is found
        header_row_idx: int, optional
            The 0-based row where this column's header is found. For better
            error messages (print row in file where an error occurred)
        """
        self.column = column
        self.header_row_idx = header_row_idx

    def __str__(self) -> str:
        if self.header_names:
            return f"Column {self.header_names[0]}"
        else:
            return "Column"

    @staticmethod
    def clean_string(string: str):
        """Make lowercase and remove separators"""
        return (
            string.replace(" ", "")
            .replace("_", "")
            .replace("-", "")
            .replace(".", "")
            .lower()
        )

    @classmethod
    def matches_header(cls, input: Union[str, None]) -> bool:
        """The given input seems to be this column's header"""
        if input is None:
            return False
        if cls.clean_string(input) in [cls.clean_string(x) for x in cls.header_names]:
            return True
        else:
            return False

    @classmethod
    def header_name(cls) -> str:
        """A header name that his header might have. For helpful error messages"""
        if cls.header_names:
            return cls.header_names[0]
        else:
            return ""

    def parameter_from_row(self, row: List[str]) -> Parameter:
        """Try to parse a Parameter instance out of a row from the grid

        Raises
        ------
        InputFileParseException
            When this row cannot be parsed into the expected Parameter type
        """
        try:
            return ParameterFactory.parse_from_key_value(
                key=self.parameter_type.field_name,
                value=row[self.column],
                parameter_types=[self.parameter_type],
            )
        except ParameterParsingError as e:
            raise InputFileParseException(e)

    def has_empty_value(self, row: List[str]) -> bool:
        """True if the given row has no value in this column"""

        return row[self.column] == "" or row[self.column] is None


class AccessionNumberColumn(ParameterColumn):

    header_names = ["accession number", "acc nr"]
    parameter_type: Type[Parameter] = AccessionNumber


class PseudonymColumn(ParameterColumn):

    header_names = ["pseudoID", "pseudonym", "name"]
    parameter_type: Type[Parameter] = PseudoName


class FolderColumn(ParameterColumn):

    header_names = ["folder", "map", "path"]
    parameter_type: Type[Parameter] = PathParameter


ALL_COLUMN_TYPES = [AccessionNumberColumn, PseudonymColumn, FolderColumn]


class TabularFile:
    """A file containing data in rows and columns

    Offers a consistent way of accessing excel files, csv files and any other formats
    """

    def rows(self) -> Iterator[List[str]]:
        """Iterates over each row in file

        Returns
        -------
        Iterator[List[str]]
            Returns list of strings for each row in file

        Raises
        ------
        InputFileException
            If anything goes wrong loading or parsing the file

        """
        raise NotImplementedError("Overwrite this method in child classes")


class ExcelFile(TabularFile):
    """An xls or xlsx file

    Some simplifying assumptions:
    * Only reads first tab, ignores any others
    * Only reads string values. Does not read any formula
    """

    def __init__(self, path: Path):
        self.path = path

    def __str__(self):
        return f"Excel file at '{self.path}'"

    def rows(self) -> Iterator[List[str]]:
        """Iterates over each row in file

        Returns
        -------
        Iterator[List[str]]
            Returns list of strings for each row in file

        Raises
        ------
        InputFileException
            If anything goes wrong loading or parsing the file

        """
        logger.info(f"Parsing '{self.path}'..")
        try:
            wb2 = load_workbook(self.path)
        except (InvalidFileException, FileNotFoundError) as e:
            raise InputFileException(f"Error reading '{self.path}':{e}")

        sheet = wb2[wb2.sheetnames[0]]  # read first sheet, ignore others

        return self.cast_rows_to_string(sheet.values)

    @staticmethod
    def cast_rows_to_string(iterator: Iterator[List]) -> Iterator[List[Optional[str]]]:
        """For standardizing data from grid-like files. Make everything string,
        except None values. Keep those None.
        """

        def str_preserve_none(input):
            if input is None:
                return None
            else:
                return str(input)

        for row in iterator:
            yield list(map(str_preserve_none, row))


class CSVFile(TabularFile):
    """A comma separated text file. Also accepts colon separated"""

    def __init__(self, path: Path):
        self.path = path

    def __str__(self):
        return f"CSVFile at '{self.path}'"

    def rows(self) -> Iterator[List[str]]:
        """Iterates over each row in file

        Notes
        -----
        Reads entire file into memory so might not work well for larger files

        Returns
        -------
        Iterator[List[str]]
            Returns list of strings for each row in file

        Raises
        ------
        InputFileException
            If anything goes wrong loading or parsing the file

        """
        logger.info(f"Parsing '{self.path}'..")
        try:
            with open(self.path, "r") as f:
                lines = f.readlines()
                rows = [
                    row for row in csv.reader(lines, dialect=sniff_dialect_safe(lines))
                ]
        except FileNotFoundError as e:
            raise InputFileException(e)

        return iter(rows)


def as_tabular_file(path: Union[str, Path]) -> TabularFile:
    """Create a TabularFile out of path, based on extension

    Parameters
    ----------
    path: Union[str, Path]
        path to tabular file

    Returns
    -------
    TabularFile
        A suitable child class

    Raises
    ------
    InputFileException
        If no suitable TabularFile class can be found for this path
    """
    path = Path(path)  # cast string to Path
    suffix = path.suffix.lower()
    if suffix in [".xls", ".xlsx"]:
        logger.debug(f"I think {path} is an Excel file")
        return ExcelFile(path)
    elif suffix in [".csv", ".txt"]:
        logger.debug(f"I think {path} is a csv file")
        return CSVFile(path)
    else:
        raise InputFileException(
            f"Unknown extension '{suffix}' I don't know how to read this file."
        )


def parse_columns(
    row: List[str], column_types: List[Type[ParameterColumn]]
) -> List[ParameterColumn]:
    """Try to find known column types in row

    Parameters
    ----------
    row: List[str]
        The row of values to parse
    column_types: List[Type[ParameterColumn]]
        The types of columns to try

    """
    columns = []

    for idx, item in enumerate(row):
        for column_type in column_types:
            if column_type.matches_header(item):
                logger.debug(
                    f"Matched '{item}' in row {row}, column {idx} to column type "
                    f"'{column_type.header_name()}'"
                )
                columns.append(column_type(column=idx))

    return columns


def find_column_headers(
    row_iterator: Iterator[List[str]], column_types: List[Type[ParameterColumn]]
) -> List[ParameterColumn]:
    """Go through each row in iterator until you find a row containing column headers

    Parameters
    ----------
    row_iterator: Iterator[List[str]]
        iterator returning lists of strings corresponding to each row of a
        grid-like file
    column_types: List[Type[ParameterColumn]]
        The types of columns to look for

    Returns
    -------
    List[ParameterColumn]
        All parameter columns found

    Raises
    ------
    InputFileException
        When no column headers can be found

    """
    logger.debug(f"Starting column header search..")

    for idx, row in enumerate(row_iterator):
        columns = parse_columns(row, column_types=column_types)
        if columns:
            logger.debug(
                f"Found {[str(x) for x in columns]} in row {idx}. "
                f"Stopping column search"
            )
            # add column header row idx for better error messages later
            for column in columns:
                column.header_row_idx = idx
            return columns
    raise InputFileException(
        f"Could not find any column headers. Looked for any"
        f" of [{','.join([x.header_name() for x in column_types])}]"
    )


def extract_parameter_grid(
    file: TabularFile,
    optional_column_types: List[Type[ParameterColumn]] = None,
    required_column_types: List[Type[ParameterColumn]] = None,
) -> JobParameterGrid:
    """Read an xls file and try to extract a grid of parameters

    Parameters
    ----------
    file: TabularFile
        Extract from this file
    optional_column_types: List[Type[ParameterColumn]], optional
        Search for these column types. Defaults to AccessionNumber and Pseudonym
    required_column_types: List[Type[ParameterColumn]], optional
        Search for these column types. Fail if not found.
        Defaults to empty list

    Raises
    ------
    InputFileException
        If parsing or reading fails for any reason

    """
    if optional_column_types is None:
        optional_column_types = [FolderColumn, AccessionNumberColumn, PseudonymColumn]
    if required_column_types is None:
        required_column_types = []

    row_iterator = file.rows()

    columns = find_column_headers(
        row_iterator, column_types=optional_column_types + required_column_types
    )
    for required in required_column_types:
        if not any(isinstance(x, required) for x in columns):
            raise InputFileParseException(
                f"Required column '{required.header_name()}' not found in file"
            )

    column_headers_idx = columns[0].header_row_idx

    # column headers found. Now build a grid one row at a time
    grid = []
    for idx, row in enumerate(row_iterator):
        try:
            grid.append(parse_row(row, columns=columns))
        except EmptyRow:
            continue  # skip empty row, try next
        except RowParseException as e:
            # Add row number (+2 as excel rows start at 1 and both idxs are 0-based)
            raise InputFileParseException(
                f"Exception in row {column_headers_idx+idx+2}: {e}"
            )

    return JobParameterGrid(grid)


def parse_row(row: List[str], columns: List[ParameterColumn]) -> List[Parameter]:
    """Is this row fit for parsing? Are there missing parameters? empty values?

    Parameters
    ----------
    row: List[str]
        String value of each cell in this row
    columns: List[ParameterColumn])
        The column names for each value in row

    Returns
    -------
    List[Parameter]
        Each value in row parsed according to column type

    Raises
    ------
    EmptyRow
        When row is empty
    PartiallyEmptyRow
        When some columns are filled but not all
    InputFileParseException
        When any of the values in row cannot be parsed according to their columns
    """
    filled = []
    empty = []
    for column in columns:
        if column.has_empty_value(row):
            empty.append(column)
        else:
            filled.append(column)

    if not filled:
        raise EmptyRow()
    if filled and empty:
        raise RowParseException(
            f"Problem in row {row}. Columns[{[str(x) for x in filled]}] have a value,"
            f" but columns [{[str(x) for x in empty]}] are empty. What do you want?"
        )
    return [column.parameter_from_row(row) for column in columns]


class InputFileException(AnonAPIException):
    pass


class InputFileParseException(InputFileException):
    pass


class EmptyRow(AnonAPIException):
    pass


class RowParseException(AnonAPIException):
    pass
